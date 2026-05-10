# mypy: disable-error-code=import-untyped
"""Export source-agnostic financial models to workbooks and DataFrames."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from waccy.core.models import FinancialStatement, ThreeStatementModel

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
CHECK_FILL = PatternFill("solid", fgColor="FCE4D6")
CURRENCY_FORMAT = '$#,##0;[Red]($#,##0);"-"'


class SheetExporter:
    """Export financial models to three-sheet XLSX workbooks."""

    def export(self, model: ThreeStatementModel, output_path: str) -> None:
        """Export a three-statement model to an XLSX workbook."""
        workbook = Workbook()
        workbook.remove(workbook.active)

        for statement in (
            model.income_statement,
            model.balance_sheet,
            model.cash_flow_statement,
        ):
            worksheet = workbook.create_sheet(statement.name)
            self._write_statement(worksheet, statement)

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)

    def _write_statement(self, worksheet: Worksheet, statement: FinancialStatement) -> None:
        worksheet.cell(row=1, column=1, value="Line Item")
        for column, period in enumerate(statement.periods, start=2):
            worksheet.cell(row=1, column=column, value=period.label)

        for row, line in enumerate(statement.lines, start=2):
            label_cell = worksheet.cell(row=row, column=1, value=line.label)
            if line.is_subtotal or line.is_check:
                label_cell.font = Font(bold=True)
            if line.is_check:
                label_cell.fill = CHECK_FILL

            for column, period in enumerate(statement.periods, start=2):
                value = line.values.get(period.label, Decimal("0"))
                value_cell = worksheet.cell(row=row, column=column, value=float(value))
                value_cell.number_format = CURRENCY_FORMAT
                if line.is_subtotal or line.is_check:
                    value_cell.font = Font(bold=True)
                if line.is_check:
                    value_cell.fill = CHECK_FILL

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL

        worksheet.freeze_panes = "B2"
        worksheet.column_dimensions["A"].width = 32
        for column in range(2, len(statement.periods) + 2):
            worksheet.column_dimensions[get_column_letter(column)].width = 14


class PandasExporter:
    """Export financial models to pandas DataFrames for follow-on modeling."""

    def export(
        self,
        model: ThreeStatementModel,
        *,
        value_type: Literal["decimal", "float"] = "decimal",
    ) -> dict[str, Any]:
        """Return one DataFrame per financial statement."""
        return {
            model.income_statement.name: self._statement_to_frame(
                model.income_statement, value_type=value_type
            ),
            model.balance_sheet.name: self._statement_to_frame(
                model.balance_sheet, value_type=value_type
            ),
            model.cash_flow_statement.name: self._statement_to_frame(
                model.cash_flow_statement, value_type=value_type
            ),
        }

    def _statement_to_frame(
        self,
        statement: FinancialStatement,
        *,
        value_type: Literal["decimal", "float"],
    ) -> Any:
        pd = self._pandas()
        rows: list[dict[str, Any]] = []
        for line in statement.lines:
            row: dict[str, Any] = {
                "line_item": line.label,
                "account_id": line.account_id,
                "is_subtotal": line.is_subtotal,
                "is_check": line.is_check,
                "source_account_ids": tuple(line.source_account_ids),
            }
            for period in statement.periods:
                value = line.values.get(period.label, Decimal("0"))
                row[period.label] = float(value) if value_type == "float" else value
            rows.append(row)
        return pd.DataFrame(rows)

    @staticmethod
    def _pandas() -> Any:
        import pandas as pd  # noqa: PLC0415

        return pd
