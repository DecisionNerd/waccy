"""Fixture-first v0.1.0 financial model pipeline tests."""

from __future__ import annotations

import importlib.util
from decimal import Decimal
from pathlib import Path
from types import ModuleType
from typing import Any

import pytest
from openpyxl import load_workbook

from tests.fixtures.sample_data import sample_edgar_fixture, sample_periods, sample_qbo_fixture
from waccy.core.models import (
    ExtractedData,
    IssueSeverity,
    MappedFinancialDataset,
    MappingOverride,
    MappingStatus,
    NormalizedFinancialDataset,
    SourceRecord,
    SourceReference,
    ValidatedFinancialDataset,
)
from waccy.core.ontology import StandardChartOfAccounts
from waccy.core.validation import validate_mapped_dataset
from waccy.extraction.mapper import DataMapper, source_record_from_dict
from waccy.modeling.builder import ModelBuilder
from waccy.modeling.exporters import SheetExporter

ROOT = Path(__file__).resolve().parents[2]


def test_core_dataset_models_round_trip() -> None:
    """Core data models carry normalized, mapped, and validated layers."""
    period = sample_periods()[0]
    source = SourceRecord(
        source_account_id="sales",
        source_account_name="Sales",
        amount=Decimal("100"),
        period_label=period.label,
        source=SourceReference(source_system="qbo", source_id="sales", source_label="Sales"),
    )
    normalized = NormalizedFinancialDataset(
        entity_name="Fixture Co",
        periods=[period],
        records=[source],
    )
    mapped = DataMapper().map_dataset(normalized)
    validated = validate_mapped_dataset(mapped)

    assert isinstance(mapped, MappedFinancialDataset)
    assert isinstance(validated, ValidatedFinancialDataset)
    assert validated.is_valid
    assert mapped.records[0].status == MappingStatus.MAPPED


def test_ontology_contains_required_accounts_and_aliases() -> None:
    """The v0.1.0 ontology includes model lines and source aliases."""
    ontology = StandardChartOfAccounts()
    required = {
        "revenue",
        "cogs",
        "operating_expenses",
        "depreciation_amortization",
        "interest_expense",
        "tax_expense",
        "cash",
        "accounts_receivable",
        "inventory",
        "ppe",
        "accumulated_depreciation",
        "accounts_payable",
        "accrued_expenses",
        "debt",
        "equity",
        "retained_earnings",
        "net_income",
        "depreciation_addback",
        "working_capital_movement",
        "capex",
        "financing_movement",
    }

    assert required.issubset(ontology.accounts)
    assert ontology.map_account("Sales", "qbo").id == "revenue"
    assert ontology.map_account("us-gaap:Revenues", "edgar").id == "revenue"
    assert ontology.map_account("us-gaap:CashAndCashEquivalentsAtCarryingValue", "edgar").id == "cash"


def test_mapper_statuses_for_mapped_ambiguous_overridden_and_unmapped() -> None:
    """The mapper emits all review states required by the API-first workflow."""
    ontology = StandardChartOfAccounts()
    ontology._aliases["ambiguous"] = ["revenue", "cogs"]
    mapper = DataMapper(ontology)
    period = sample_periods()[0]
    dataset = NormalizedFinancialDataset(
        entity_name="Fixture Co",
        periods=[period],
        records=[
            _source_record("sales", "Sales"),
            _source_record("ambiguous", "ambiguous"),
            _source_record("manual", "Manual Account"),
            _source_record("mystery", "Mystery Account"),
        ],
    )

    mapped = mapper.map_dataset(
        dataset,
        overrides={"manual": MappingOverride(account_id="cash", note="Known bank account")},
    )
    statuses = {record.source_record.source_account_id: record.status for record in mapped.records}

    assert statuses == {
        "sales": MappingStatus.MAPPED,
        "ambiguous": MappingStatus.AMBIGUOUS,
        "manual": MappingStatus.OVERRIDDEN,
        "mystery": MappingStatus.UNMAPPED,
    }


def test_qbo_fixture_extractor_produces_normalized_dataset() -> None:
    """The QBO extractor accepts fixture/dict input and normalizes it."""
    extractor_class = _load_extension_class(
        ROOT / "extensions" / "waccy-quickbooks" / "src" / "waccy_quickbooks" / "extractor.py",
        "local_waccy_quickbooks_extractor",
        "QuickBooksExtractor",
    )

    extracted = extractor_class().extract({"fixture": sample_qbo_fixture()})
    normalized = DataMapper().normalize(extracted)

    assert normalized.entity_name == "Fixture Co"
    assert len(normalized.periods) == 2
    assert len(normalized.records) == len(sample_qbo_fixture()["records"])
    assert {record.source.source_system for record in normalized.records} == {"qbo"}


def test_qbo_fixture_extractor_rejects_invalid_fixture_shapes() -> None:
    """The QBO extractor returns clear errors for invalid fixture shapes."""
    extractor_class = _load_extension_class(
        ROOT / "extensions" / "waccy-quickbooks" / "src" / "waccy_quickbooks" / "extractor.py",
        "local_waccy_quickbooks_extractor_invalid",
        "QuickBooksExtractor",
    )
    extractor = extractor_class()

    with pytest.raises(ValueError, match="dictionary payload"):
        extractor.extract({"fixture": "not-a-dict"})
    with pytest.raises(ValueError, match="'records' list"):
        extractor.extract({"fixture": {}, "records": []})
    with pytest.raises(ValueError, match="'records' list"):
        extractor.extract({"fixture": {"records": "not-a-list"}})
    with pytest.raises(ValueError, match="records must be dictionaries"):
        extractor.extract({"fixture": {"records": ["not-a-dict"]}})
    with pytest.raises(ValueError, match="periods must be dictionaries"):
        extractor.extract({"fixture": {"records": [], "periods": ["not-a-dict"]}})
    with pytest.raises(ValueError, match="missing required keys: end_date"):
        extractor.extract({"fixture": {"records": [], "periods": [{"label": "2024", "start_date": "2024-01-01"}]}})
    with pytest.raises(ValueError, match="accounts must be a list"):
        extractor.extract({"fixture": {"records": [], "accounts": "not-a-list"}})
    with pytest.raises(ValueError, match="accounts must be dictionaries"):
        extractor.extract({"fixture": {"records": [], "accounts": ["not-a-dict"]}})


def test_edgar_fixture_extractor_produces_normalized_dataset() -> None:
    """The EDGAR extractor accepts XBRL-like fixture/dict input and normalizes it."""
    extractor_class = _load_extension_class(
        ROOT / "extensions" / "waccy-edgar" / "src" / "waccy_edgar" / "extractor.py",
        "local_waccy_edgar_extractor",
        "EdgarExtractor",
    )

    extracted = extractor_class().extract({"fixture": sample_edgar_fixture()})
    normalized = DataMapper().normalize(extracted)

    assert normalized.entity_name == "Fixture Co"
    assert len(normalized.periods) == 2
    assert normalized.records[0].source.source_system == "edgar"


def test_edgar_fixture_extractor_rejects_invalid_fixture_shapes() -> None:
    """The EDGAR extractor returns clear errors for invalid fixture shapes."""
    extractor_class = _load_extension_class(
        ROOT / "extensions" / "waccy-edgar" / "src" / "waccy_edgar" / "extractor.py",
        "local_waccy_edgar_extractor_invalid",
        "EdgarExtractor",
    )
    extractor = extractor_class()

    with pytest.raises(ValueError, match="dictionary payload"):
        extractor.extract({"fixture": "not-a-dict"})
    with pytest.raises(ValueError, match="'records' list"):
        extractor.extract({"fixture": {}, "records": []})
    with pytest.raises(ValueError, match="'records' list"):
        extractor.extract({"fixture": {"records": "not-a-list"}})
    with pytest.raises(ValueError, match="records must be dictionaries"):
        extractor.extract({"fixture": {"records": ["not-a-dict"]}})
    with pytest.raises(ValueError, match="periods must be dictionaries"):
        extractor.extract({"fixture": {"records": [], "periods": ["not-a-dict"]}})
    with pytest.raises(ValueError, match="missing required keys: end_date"):
        extractor.extract({"fixture": {"records": [], "periods": [{"label": "2024", "start_date": "2024-01-01"}]}})


def test_validation_reports_unmapped_missing_periods_and_duplicate_periods() -> None:
    """Validation catches unmapped records and period defects."""
    period = sample_periods()[0]
    source = _source_record("mystery", "Mystery Account")
    dataset = NormalizedFinancialDataset(
        entity_name="Fixture Co",
        periods=[period, period],
        records=[source_record_from_dict({"name": "Sales", "period": "2025", "amount": 1}, "qbo"), source],
    )
    mapped = DataMapper().map_dataset(dataset)
    validated = validate_mapped_dataset(mapped)
    issue_codes = {issue.code for issue in validated.issues}

    assert {"duplicate_period", "missing_period", "unmapped_account"}.issubset(issue_codes)
    assert not validated.is_valid


def test_mapper_infers_quarterly_and_monthly_reporting_periods() -> None:
    """Mapper-created periods preserve common monthly and quarterly labels."""
    extracted = ExtractedData(
        entity_name="Fixture Co",
        source_records=[
            source_record_from_dict({"name": "Sales", "period": "2024Q1", "amount": 1}, "qbo"),
            source_record_from_dict({"name": "Sales", "period": "2024-02", "amount": 1}, "qbo"),
            source_record_from_dict({"name": "Sales", "period": "202403", "amount": 1}, "qbo"),
        ],
        metadata={"source": "qbo"},
    )

    normalized = DataMapper().normalize(extracted)
    periods = {period.label: period for period in normalized.periods}

    assert periods["2024Q1"].start_date.isoformat() == "2024-01-01"
    assert periods["2024Q1"].end_date.isoformat() == "2024-03-31"
    assert periods["2024-02"].start_date.isoformat() == "2024-02-01"
    assert periods["202403"].end_date.isoformat() == "2024-03-31"


def test_model_builder_reports_balance_and_cash_flow_issues() -> None:
    """Model validation includes balance-sheet and cash-flow checks."""
    fixture = sample_qbo_fixture()
    fixture["records"] = [
        *fixture["records"],
        {"name": "Checking", "period": "2024", "amount": 1, "statement": "balance_sheet"},
    ]
    extracted = ExtractedData(
        entity_name="Fixture Co",
        periods=sample_periods(),
        source_records=[source_record_from_dict(record, "qbo") for record in fixture["records"]],
        metadata={"source": "qbo"},
    )
    model = ModelBuilder().build_three_statement_model(extracted)
    issue_codes = {issue.code for issue in model.validation_issues}

    assert "balance_sheet_imbalance" in issue_codes
    assert "cash_flow_tie_out_failure" in issue_codes


def test_model_builder_builds_qbo_and_edgar_three_statement_models() -> None:
    """QBO and EDGAR fixtures both build the source-agnostic model."""
    builder = ModelBuilder()
    qbo_model = builder.build_three_statement_model(_extracted(sample_qbo_fixture(), "qbo"))
    edgar_model = builder.build_three_statement_model(_extracted(sample_edgar_fixture(), "edgar"))

    assert _line_value(qbo_model.income_statement.lines, "Net Income", "2024") == Decimal("316")
    assert _line_value(qbo_model.balance_sheet.lines, "Balance Check", "2024") == Decimal("0")
    assert _line_value(qbo_model.cash_flow_statement.lines, "Cash Flow Tie-Out", "2024") == Decimal("0")
    assert _line_value(edgar_model.income_statement.lines, "Revenue", "2024") == Decimal("1200")
    assert all(issue.severity != IssueSeverity.ERROR for issue in qbo_model.validation_issues)
    assert all(issue.severity != IssueSeverity.ERROR for issue in edgar_model.validation_issues)


def test_xlsx_export_writes_three_sheet_workbook(tmp_path: Path) -> None:
    """The XLSX exporter writes three formatted statement sheets."""
    model = ModelBuilder().build_three_statement_model(_extracted(sample_qbo_fixture(), "qbo"))
    output_path = tmp_path / "fixture-model.xlsx"

    SheetExporter().export(model, str(output_path))
    workbook = load_workbook(output_path)
    try:
        assert workbook.sheetnames == ["Income Statement", "Balance Sheet", "Cash Flow Statement"]
        income_statement = workbook["Income Statement"]
        balance_sheet = workbook["Balance Sheet"]
        cash_flow_statement = workbook["Cash Flow Statement"]
        assert income_statement["B1"].value == "2023"
        assert income_statement["C11"].value == 316
        assert balance_sheet["A15"].value == "Balance Check"
        assert balance_sheet["C15"].value == 0
        assert cash_flow_statement["A8"].value == "Cash Flow Tie-Out"
        assert cash_flow_statement["C8"].value == 0
    finally:
        workbook.close()


def _source_record(account_id: str, account_name: str) -> SourceRecord:
    return SourceRecord(
        source_account_id=account_id,
        source_account_name=account_name,
        amount=Decimal("100"),
        period_label="2023",
        source=SourceReference(source_system="fixture", source_id=account_id, source_label=account_name),
    )


def _extracted(fixture: dict[str, object], source_system: str) -> ExtractedData:
    raw_records = fixture["records"]
    if not isinstance(raw_records, list):
        raise AssertionError(f"Fixture {fixture['entity_name']!r} records must be a list.")
    for record in raw_records:
        if not isinstance(record, dict):
            raise AssertionError(
                f"Fixture {fixture['entity_name']!r} records must be dictionaries: {record!r}"
            )

    return ExtractedData(
        entity_name=str(fixture["entity_name"]),
        periods=sample_periods(),
        source_records=[source_record_from_dict(record, source_system) for record in raw_records],
        metadata={"source": source_system},
    )


def _line_value(lines: list[object], label: str, period: str) -> Decimal:
    for line in lines:
        if line.label == label:
            if period not in line.values:
                raise AssertionError(f"Line {label!r} is missing period {period!r}")
            return line.values[period]
    raise AssertionError(f"Missing line {label!r}")


def _load_extension_class(path: Path, module_name: str, class_name: str) -> type[Any]:
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise AssertionError(f"Could not load extension module from {path}")
    module = importlib.util.module_from_spec(spec)
    if not isinstance(module, ModuleType):
        raise AssertionError(f"Could not create extension module from {path}")
    spec.loader.exec_module(module)
    loaded_class = getattr(module, class_name)
    if not isinstance(loaded_class, type):
        raise AssertionError(f"{class_name} is not a class")
    return loaded_class
